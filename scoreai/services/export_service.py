"""
エクスポート機能を提供するサービス層
"""
import csv
import logging
from typing import List, Dict, Any, Optional
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime

logger = logging.getLogger(__name__)

# オプショナルな依存関係のチェック
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl is not installed. Excel export will not be available.")

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab is not installed. PDF export will not be available.")


class ExportService:
    """エクスポート機能を提供するサービスクラス"""
    
    @staticmethod
    def export_to_csv(
        headers: List[str],
        data: List[List[Any]],
        filename: str,
        encoding: str = 'utf-8-sig'
    ) -> HttpResponse:
        """
        CSV形式でエクスポート
        
        Args:
            headers: ヘッダー行のリスト
            data: データ行のリスト（各行は値のリスト）
            filename: ファイル名
            encoding: エンコーディング（デフォルト: utf-8-sig）
            
        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv; charset=' + encoding)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # BOMを追加してExcelで正しく開けるようにする（UTF-8の場合）
        if encoding == 'utf-8-sig':
            response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for row in data:
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_to_excel(
        headers: List[str],
        data: List[List[Any]],
        filename: str,
        sheet_name: str = 'Sheet1',
        title: Optional[str] = None
    ) -> HttpResponse:
        """
        Excel形式でエクスポート
        
        Args:
            headers: ヘッダー行のリスト
            data: データ行のリスト（各行は値のリスト）
            filename: ファイル名
            sheet_name: シート名
            title: タイトル（オプション）
            
        Returns:
            HttpResponse with Excel content
            
        Raises:
            ImportError: openpyxlがインストールされていない場合
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # スタイルの定義
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        
        # タイトルの追加
        if title:
            ws.merge_cells(f'A{row_num}:{get_column_letter(len(headers))}{row_num}')
            title_cell = ws.cell(row=row_num, column=1)
            title_cell.value = title
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
            ws.row_dimensions[1].height = 25
        
        # ヘッダー行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        row_num += 1
        
        # データ行
        for row_data in data:
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
            row_num += 1
        
        # 列幅の自動調整
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # レスポンスの作成
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_to_pdf(
        title: str,
        headers: List[str],
        data: List[List[Any]],
        filename: str,
        additional_info: Optional[Dict[str, Any]] = None
    ) -> HttpResponse:
        """
        PDF形式でエクスポート
        
        Args:
            title: レポートのタイトル
            headers: ヘッダー行のリスト
            data: データ行のリスト（各行は値のリスト）
            filename: ファイル名
            additional_info: 追加情報（オプション）
            
        Returns:
            HttpResponse with PDF content
            
        Raises:
            ImportError: reportlabがインストールされていない場合
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is not installed. Please install it: pip install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # スタイルの定義
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # 中央揃え
        )
        
        # タイトル
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # 追加情報の表示
        if additional_info:
            info_text = []
            for key, value in additional_info.items():
                info_text.append(f"<b>{key}:</b> {value}")
            story.append(Paragraph("<br/>".join(info_text), styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
        
        # テーブルの作成
        table_data = [headers] + data
        
        # テーブルのスタイル
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(table_style)
        story.append(table)
        
        # フッター（生成日時）
        story.append(Spacer(1, 0.3 * inch))
        footer_text = f"生成日時: {timezone.now().strftime('%Y年%m月%d日 %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # PDFの生成
        doc.build(story)
        
        # レスポンスの作成
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        
        return response

