"""
財務会議資料生成サービス

CSVファイルから財務会議用のExcel資料を自動生成する。
- 部門別損益計算書（PL）
- 月次推移損益計算書
- 貸借対照表（BS）
- エグゼクティブサマリー
"""
import logging
from dataclasses import dataclass, field
from typing import Optional, Dict, Any, Union, BinaryIO, List
from io import BytesIO, StringIO
from decimal import Decimal
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins, PrintPageSetup

from ..utils.csv_utils import detect_encoding

logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """レポート設定"""
    company_name: str
    target_month: int
    target_year: int
    target_f_rate: float = 0.30      # 原価率目標
    target_l_rate: float = 0.30      # 人件費率目標
    target_fl_rate: float = 0.60    # FL率目標
    target_current_ratio: float = 200.0   # 流動比率目標
    target_equity_ratio: float = 30.0     # 自己資本比率目標


class FinancialReportGenerator:
    """財務会議資料ジェネレーター"""
    
    # 色定義
    COLORS = {
        'header_bg': 'E6F0FF',      # ヘッダー背景（薄い青）
        'header_font': '1F4E79',    # ヘッダー文字色
        'section_bg': 'F2F2F2',     # セクション背景（薄いグレー）
        'total_bg': 'DCE6F1',       # 合計行背景
        'highlight_good': 'C6EFCE', # 良好（緑）
        'highlight_bad': 'FFC7CE',  # 注意（赤）
        'border_color': 'B4B4B4',   # 罫線色
    }
    
    def __init__(self, config: ReportConfig):
        """
        Args:
            config: レポート設定
        """
        self.config = config
        self.pl_bumon_df: Optional[pd.DataFrame] = None
        self.pl_bumon_zenki_df: Optional[pd.DataFrame] = None
        self.pl_suii_df: Optional[pd.DataFrame] = None
        self.pl_suii_zenki_df: Optional[pd.DataFrame] = None
        self.bs_df: Optional[pd.DataFrame] = None
        
        # スタイル定義
        self._init_styles()
    
    def _init_styles(self):
        """スタイルを初期化"""
        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border_color']),
            right=Side(style='thin', color=self.COLORS['border_color']),
            top=Side(style='thin', color=self.COLORS['border_color']),
            bottom=Side(style='thin', color=self.COLORS['border_color'])
        )
        
        self.header_fill = PatternFill(
            start_color=self.COLORS['header_bg'],
            end_color=self.COLORS['header_bg'],
            fill_type='solid'
        )
        
        self.header_font = Font(
            bold=True,
            size=10,
            color=self.COLORS['header_font']
        )
        
        self.section_fill = PatternFill(
            start_color=self.COLORS['section_bg'],
            end_color=self.COLORS['section_bg'],
            fill_type='solid'
        )
        
        self.total_fill = PatternFill(
            start_color=self.COLORS['total_bg'],
            end_color=self.COLORS['total_bg'],
            fill_type='solid'
        )
    
    def load_data(
        self,
        pl_bumon_file: Optional[Union[str, bytes, BinaryIO]] = None,
        pl_bumon_zenki_file: Optional[Union[str, bytes, BinaryIO]] = None,
        pl_suii_file: Optional[Union[str, bytes, BinaryIO]] = None,
        pl_suii_zenki_file: Optional[Union[str, bytes, BinaryIO]] = None,
        bs_file: Optional[Union[str, bytes, BinaryIO]] = None,
    ) -> None:
        """
        CSVファイルを読み込む
        
        Args:
            pl_bumon_file: 部門別PL（当期）
            pl_bumon_zenki_file: 部門別PL（前期）
            pl_suii_file: PL月次推移（当期）
            pl_suii_zenki_file: PL月次推移（前期）
            bs_file: 貸借対照表
        """
        if pl_bumon_file:
            self.pl_bumon_df = self._read_csv(pl_bumon_file)
            logger.info(f"部門別PL（当期）を読み込みました: {self.pl_bumon_df.shape}")
        
        if pl_bumon_zenki_file:
            self.pl_bumon_zenki_df = self._read_csv(pl_bumon_zenki_file)
            logger.info(f"部門別PL（前期）を読み込みました: {self.pl_bumon_zenki_df.shape}")
        
        if pl_suii_file:
            self.pl_suii_df = self._read_csv(pl_suii_file)
            logger.info(f"PL月次推移（当期）を読み込みました: {self.pl_suii_df.shape}")
        
        if pl_suii_zenki_file:
            self.pl_suii_zenki_df = self._read_csv(pl_suii_zenki_file)
            logger.info(f"PL月次推移（前期）を読み込みました: {self.pl_suii_zenki_df.shape}")
        
        if bs_file:
            self.bs_df = self._read_csv(bs_file)
            logger.info(f"貸借対照表を読み込みました: {self.bs_df.shape}")
    
    def _read_csv(self, file: Union[str, bytes, BinaryIO]) -> pd.DataFrame:
        """
        CSVファイルを読み込んでDataFrameに変換
        
        Args:
            file: ファイルパス、バイト列、またはファイルオブジェクト
            
        Returns:
            pandas.DataFrame
        """
        # ファイルパスの場合
        if isinstance(file, str):
            with open(file, 'rb') as f:
                content = f.read()
        # バイト列の場合
        elif isinstance(file, bytes):
            content = file
        # ファイルオブジェクトの場合
        else:
            content = file.read()
            if hasattr(file, 'seek'):
                file.seek(0)
        
        # エンコーディング検出
        encoding = detect_encoding(content)
        
        # 複数のエンコーディングを試す
        encodings_to_try = [encoding, 'shift-jis', 'utf-8', 'cp932', 'euc-jp']
        
        for enc in encodings_to_try:
            try:
                decoded = content.decode(enc)
                df = pd.read_csv(StringIO(decoded), header=0)
                logger.debug(f"CSV読み込み成功: encoding={enc}")
                logger.debug(f"CSV列名: {list(df.columns)}")
                logger.debug(f"CSVデータ行数: {len(df)}")
                if len(df) > 0:
                    logger.debug(f"最初の5行の勘定科目: {df.iloc[:5, 0].tolist()}")
                return df
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as e:
                logger.warning(f"CSV読み込みエラー (encoding={enc}): {e}")
                continue
        
        raise ValueError("CSVファイルのエンコーディングを検出できませんでした。")
    
    def generate(self) -> BytesIO:
        """
        Excel財務会議資料を生成
        
        Returns:
            BytesIO: 生成されたExcelファイル
        """
        wb = Workbook()
        
        # デフォルトシートを削除
        default_sheet = wb.active
        
        # 各シートを生成
        self._build_executive_summary(wb)
        self._build_pl_bumon(wb)
        self._build_pl_suii(wb)
        self._build_bs(wb)
        
        # デフォルトシートを削除（他のシートがある場合）
        if len(wb.sheetnames) > 1:
            del wb[default_sheet.title]
        
        # BytesIOに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _build_executive_summary(self, wb: Workbook) -> None:
        """エグゼクティブサマリーシートを作成"""
        ws = wb.create_sheet(title="エグゼクティブサマリー")
        
        # ページ設定（A4縦）
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        
        row = 1
        
        # タイトル
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{self.config.company_name} 財務会議資料"
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # サブタイトル
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{self.config.target_year}年{self.config.target_month}月度"
        cell.font = Font(size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 2
        
        # 経営指標セクション
        row = self._add_summary_section(ws, row, "売上・利益", [
            ("当月売上高", self._get_current_month_sales(), "円"),
            ("累計売上高", self._get_cumulative_sales(), "円"),
            ("当月粗利益", self._get_current_month_gross_profit(), "円"),
            ("累計粗利益", self._get_cumulative_gross_profit(), "円"),
            ("当月営業利益", self._get_current_month_operating_profit(), "円"),
            ("累計営業利益", self._get_cumulative_operating_profit(), "円"),
        ])
        row += 1
        
        # 利益率セクション
        row = self._add_summary_section(ws, row, "利益率", [
            ("粗利益率", self._get_gross_profit_rate(), "%"),
            ("営業利益率", self._get_operating_profit_rate(), "%"),
            ("原価率（F率）", self._get_f_rate(), "%", self.config.target_f_rate * 100),
            ("人件費率（L率）", self._get_l_rate(), "%", self.config.target_l_rate * 100),
            ("FL率", self._get_fl_rate(), "%", self.config.target_fl_rate * 100),
        ])
        row += 1
        
        # 財務指標セクション
        row = self._add_summary_section(ws, row, "財務指標", [
            ("流動比率", self._get_current_ratio(), "%", self.config.target_current_ratio),
            ("自己資本比率", self._get_equity_ratio(), "%", self.config.target_equity_ratio),
        ])
        
        # 列幅設定
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
    
    def _add_summary_section(
        self,
        ws,
        start_row: int,
        section_title: str,
        items: List[tuple]
    ) -> int:
        """サマリーセクションを追加"""
        row = start_row
        
        # セクションタイトル
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"■ {section_title}"
        cell.font = Font(bold=True, size=11)
        cell.fill = self.section_fill
        ws.row_dimensions[row].height = 22
        row += 1
        
        # ヘッダー
        headers = ['', '項目', '実績', '単位', '目標', '達成率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # データ行
        for item in items:
            name = item[0]
            value = item[1]
            unit = item[2]
            target = item[3] if len(item) > 3 else None
            
            ws.cell(row=row, column=2).value = name
            ws.cell(row=row, column=2).border = self.thin_border
            
            # 値のフォーマット
            if value is not None:
                if unit == "円":
                    ws.cell(row=row, column=3).value = value
                    ws.cell(row=row, column=3).number_format = '#,##0'
                else:
                    ws.cell(row=row, column=3).value = value
                    ws.cell(row=row, column=3).number_format = '0.0'
            else:
                ws.cell(row=row, column=3).value = "-"
            ws.cell(row=row, column=3).border = self.thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            
            ws.cell(row=row, column=4).value = unit
            ws.cell(row=row, column=4).border = self.thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            
            # 目標
            if target is not None:
                ws.cell(row=row, column=5).value = target
                ws.cell(row=row, column=5).number_format = '0.0'
            else:
                ws.cell(row=row, column=5).value = "-"
            ws.cell(row=row, column=5).border = self.thin_border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            
            # 達成率
            if target is not None and value is not None and target != 0:
                achievement = (value / target) * 100
                ws.cell(row=row, column=6).value = f"{achievement:.1f}%"
                # 色分け（原価率系は逆転）
                if name in ['原価率（F率）', '人件費率（L率）', 'FL率']:
                    if achievement <= 100:
                        ws.cell(row=row, column=6).fill = PatternFill(
                            start_color=self.COLORS['highlight_good'],
                            end_color=self.COLORS['highlight_good'],
                            fill_type='solid'
                        )
                    else:
                        ws.cell(row=row, column=6).fill = PatternFill(
                            start_color=self.COLORS['highlight_bad'],
                            end_color=self.COLORS['highlight_bad'],
                            fill_type='solid'
                        )
                else:
                    if achievement >= 100:
                        ws.cell(row=row, column=6).fill = PatternFill(
                            start_color=self.COLORS['highlight_good'],
                            end_color=self.COLORS['highlight_good'],
                            fill_type='solid'
                        )
                    elif achievement < 80:
                        ws.cell(row=row, column=6).fill = PatternFill(
                            start_color=self.COLORS['highlight_bad'],
                            end_color=self.COLORS['highlight_bad'],
                            fill_type='solid'
                        )
            else:
                ws.cell(row=row, column=6).value = "-"
            ws.cell(row=row, column=6).border = self.thin_border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            
            row += 1
        
        return row
    
    def _build_pl_bumon(self, wb: Workbook) -> None:
        """部門別PLシートを作成（前期比較付き）"""
        if self.pl_bumon_df is None:
            logger.warning("部門別PLデータがありません")
            return
        
        ws = wb.create_sheet(title=f"部門別PL（{self.config.target_month}月）")
        
        # ページ設定（A3横）
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_title_cols = 'A:A'  # A列を繰り返し印刷
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)
        
        # タイトル
        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{self.config.company_name} 部門別損益計算書"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # サブタイトル
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"対象期間: {self.config.target_year}年{self.config.target_month}月"
        cell.font = Font(size=10)
        row += 2
        
        # データの準備
        df = self.pl_bumon_df.copy()
        df_zenki = self.pl_bumon_zenki_df.copy() if self.pl_bumon_zenki_df is not None else None
        
        # CSVの構造: 列0=セクション名、列1=勘定科目名、列2以降=データ
        # 勘定科目名は列0が空なら列1を使用、そうでなければ列0を使用
        section_col = df.columns[0] if len(df.columns) > 0 else None
        detail_col = df.columns[1] if len(df.columns) > 1 else None
        if section_col is None:
            return
        
        # 前期も同様
        section_col_zenki = df_zenki.columns[0] if df_zenki is not None and len(df_zenki.columns) > 0 else None
        detail_col_zenki = df_zenki.columns[1] if df_zenki is not None and len(df_zenki.columns) > 1 else None
        
        # 部門列を取得（列0,1以外、かつ除外キーワードを含まない列）
        exclude_keywords = ['勘定科目', '科目', '科目名', '勘定', 'account', 'unnamed']
        department_cols = [
            col for col in df.columns[2:]  # 列2以降からデータ列
            if not pd.isna(col) and str(col).strip() 
            and not any(kw in str(col).lower() for kw in exclude_keywords)
        ]
        
        logger.debug(f"部門別PL 部門列: {department_cols}")
        
        # 勘定科目名を取得するヘルパー関数
        def get_account_name(row_data, sec_col, det_col):
            sec_val = str(row_data[sec_col]).strip() if pd.notna(row_data[sec_col]) else ""
            det_val = str(row_data[det_col]).strip() if det_col and pd.notna(row_data[det_col]) else ""
            # セクション列に値があればそれを使用、なければ詳細列を使用
            return sec_val if sec_val else det_val
        
        # 売上高合計行を探す（比率計算用）
        sales_row = None
        for idx, row_data in df.iterrows():
            account_name = get_account_name(row_data, section_col, detail_col)
            if '売上高合計' in account_name or account_name == '売上高':
                sales_row = row_data
                break
        
        # ヘッダー行1: 部門名
        col = 1
        ws.cell(row=row, column=col).value = "勘定科目"
        ws.cell(row=row, column=col).font = self.header_font
        ws.cell(row=row, column=col).fill = self.header_fill
        ws.cell(row=row, column=col).border = self.thin_border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        col += 1
        
        for dept in department_cols:
            # 部門名を4列にマージ
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
            ws.cell(row=row, column=col).value = str(dept)
            ws.cell(row=row, column=col).font = self.header_font
            ws.cell(row=row, column=col).fill = self.header_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            for c in range(col, col+4):
                ws.cell(row=row, column=c).border = self.thin_border
            col += 4
        row += 1
        
        # ヘッダー行2: 金額/比率/前期/前期比
        col = 1
        ws.cell(row=row, column=col).value = ""
        ws.cell(row=row, column=col).border = self.thin_border
        col += 1
        
        sub_headers = ["金額", "比率", "前期", "前期比"]
        for dept in department_cols:
            for sh in sub_headers:
                ws.cell(row=row, column=col).value = sh
                ws.cell(row=row, column=col).font = Font(bold=True, size=9)
                ws.cell(row=row, column=col).fill = self.header_fill
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                col += 1
        row += 1
        
        # データ行
        for idx, row_data in df.iterrows():
            col = 1
            account_name = get_account_name(row_data, section_col, detail_col)
            
            # 空の行はスキップ
            if not account_name:
                continue
            
            # セクション行のスタイル（列0に値がある行）
            sec_val = str(row_data[section_col]).strip() if pd.notna(row_data[section_col]) else ""
            det_val = str(row_data[detail_col]).strip() if detail_col and pd.notna(row_data[detail_col]) else ""
            is_detail = not sec_val and det_val  # 列0が空で列1に値がある場合は詳細行
            is_section = sec_val in ['売上高', '売上原価', '販売費及び一般管理費', '営業外収益', '営業外費用', '特別利益', '特別損失', '営業外収益合計', '営業外費用合計', '特別利益合計', '特別損失合計', '当期純利益合計']
            is_total = '合計' in account_name or '利益' in account_name
            
            # 詳細行はインデントを追加
            display_name = f"  {account_name}" if is_detail else account_name
            
            # 勘定科目名
            ws.cell(row=row, column=col).value = display_name
            ws.cell(row=row, column=col).border = self.thin_border
            
            if is_section:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.section_fill
            elif is_total:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.total_fill
            
            col += 1
            
            # 各部門のデータ
            for dept in department_cols:
                # 当期金額
                current_value = row_data[dept] if pd.notna(row_data[dept]) else 0
                try:
                    current_value = float(current_value)
                except (ValueError, TypeError):
                    current_value = 0
                
                # 前期金額
                prev_value = None  # Noneは前期データなし、0は前期が0円
                if df_zenki is not None and section_col_zenki is not None and dept in df_zenki.columns:
                    # 勘定科目名で前期データを検索（行タイプも一致させる）
                    for zenki_idx, zenki_data in df_zenki.iterrows():
                        zenki_sec = str(zenki_data[section_col_zenki]).strip() if pd.notna(zenki_data[section_col_zenki]) else ""
                        zenki_det = str(zenki_data[detail_col_zenki]).strip() if detail_col_zenki and pd.notna(zenki_data[detail_col_zenki]) else ""
                        zenki_is_detail = not zenki_sec and zenki_det  # 前期の行が詳細行かどうか
                        zenki_account = zenki_det if zenki_is_detail else zenki_sec
                        
                        # 当期と前期で同じ行タイプ（詳細行 or セクション/合計行）のみマッチ
                        if zenki_account == account_name and zenki_is_detail == is_detail:
                            prev_val = zenki_data[dept]
                            try:
                                prev_value = float(prev_val) if pd.notna(prev_val) else 0
                            except (ValueError, TypeError):
                                prev_value = 0
                            break
                
                # 比率（売上高に対する比率）
                ratio = 0
                if sales_row is not None and dept in sales_row.index:
                    sales_value = sales_row[dept]
                    try:
                        sales_value = float(sales_value) if pd.notna(sales_value) else 0
                        if sales_value != 0:
                            ratio = current_value / sales_value
                    except (ValueError, TypeError):
                        pass
                
                # 前期比（前期データがある場合のみ計算）
                yoy = None
                if prev_value is not None and prev_value != 0:
                    yoy = (current_value - prev_value) / abs(prev_value)
                
                # 金額
                ws.cell(row=row, column=col).value = current_value
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 比率
                ws.cell(row=row, column=col).value = ratio
                ws.cell(row=row, column=col).number_format = '0.0%'
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 前期（データがない場合は空白）
                if prev_value is not None:
                    ws.cell(row=row, column=col).value = prev_value
                    ws.cell(row=row, column=col).number_format = '#,##0'
                else:
                    ws.cell(row=row, column=col).value = ""
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 前期比
                if yoy is not None:
                    ws.cell(row=row, column=col).value = yoy
                    ws.cell(row=row, column=col).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=col).value = ""
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
            
            row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        for i in range(2, col):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    def _build_pl_suii(self, wb: Workbook) -> None:
        """PL月次推移シートを作成（前期比較付き）"""
        if self.pl_suii_df is None:
            logger.warning("PL月次推移データがありません")
            return
        
        ws = wb.create_sheet(title="PL月次推移")
        
        # ページ設定（A3横）
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_title_cols = 'A:A'  # A列を繰り返し印刷
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)
        
        # タイトル
        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{self.config.company_name} 月次損益計算書推移"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # サブタイトル
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{self.config.target_year}年度"
        cell.font = Font(size=10)
        row += 2
        
        # データの準備
        df = self.pl_suii_df.copy()
        df_zenki = self.pl_suii_zenki_df.copy() if self.pl_suii_zenki_df is not None else None
        
        # CSVの構造: 列0=セクション名、列1=勘定科目名、列2以降=データ
        section_col = df.columns[0] if len(df.columns) > 0 else None
        detail_col = df.columns[1] if len(df.columns) > 1 else None
        if section_col is None:
            return
        
        # 前期も同様
        section_col_zenki = df_zenki.columns[0] if df_zenki is not None and len(df_zenki.columns) > 0 else None
        detail_col_zenki = df_zenki.columns[1] if df_zenki is not None and len(df_zenki.columns) > 1 else None
        
        # 月列を取得（列0,1以外、かつ除外キーワードを含まない列）
        exclude_keywords = ['勘定科目', '科目', '科目名', '勘定', 'account', 'unnamed']
        month_cols = [
            col for col in df.columns[2:]  # 列2以降からデータ列
            if not pd.isna(col) and str(col).strip()
            and not any(kw in str(col).lower() for kw in exclude_keywords)
        ]
        
        logger.debug(f"月次推移PL 月列: {month_cols}")
        
        # 勘定科目名を取得するヘルパー関数
        def get_account_name(row_data, sec_col, det_col):
            sec_val = str(row_data[sec_col]).strip() if pd.notna(row_data[sec_col]) else ""
            det_val = str(row_data[det_col]).strip() if det_col and pd.notna(row_data[det_col]) else ""
            return sec_val if sec_val else det_val
        
        # 売上高合計行を探す（比率計算用）
        sales_data = {}
        for idx, row_data in df.iterrows():
            account_name = get_account_name(row_data, section_col, detail_col)
            if '売上高合計' in account_name or account_name == '売上高':
                for month in month_cols:
                    try:
                        sales_data[month] = float(row_data[month]) if pd.notna(row_data[month]) else 0
                    except (ValueError, TypeError):
                        sales_data[month] = 0
                break
        
        # ヘッダー行1: 月名
        col = 1
        ws.cell(row=row, column=col).value = "勘定科目"
        ws.cell(row=row, column=col).font = self.header_font
        ws.cell(row=row, column=col).fill = self.header_fill
        ws.cell(row=row, column=col).border = self.thin_border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        col += 1
        
        for month in month_cols:
            # 月名を4列にマージ
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
            ws.cell(row=row, column=col).value = str(month)
            ws.cell(row=row, column=col).font = self.header_font
            ws.cell(row=row, column=col).fill = self.header_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            for c in range(col, col+4):
                ws.cell(row=row, column=c).border = self.thin_border
            col += 4
        row += 1
        
        # ヘッダー行2: 金額/比率/前期/前期比
        col = 1
        ws.cell(row=row, column=col).value = ""
        ws.cell(row=row, column=col).border = self.thin_border
        col += 1
        
        sub_headers = ["金額", "比率", "前期", "前期比"]
        for month in month_cols:
            for sh in sub_headers:
                ws.cell(row=row, column=col).value = sh
                ws.cell(row=row, column=col).font = Font(bold=True, size=9)
                ws.cell(row=row, column=col).fill = self.header_fill
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                col += 1
        row += 1
        
        # データ行
        for idx, row_data in df.iterrows():
            col = 1
            account_name = get_account_name(row_data, section_col, detail_col)
            
            # 空の行はスキップ
            if not account_name:
                continue
            
            # セクション行のスタイル（列0に値がある行）
            sec_val = str(row_data[section_col]).strip() if pd.notna(row_data[section_col]) else ""
            det_val = str(row_data[detail_col]).strip() if detail_col and pd.notna(row_data[detail_col]) else ""
            is_detail = not sec_val and det_val  # 列0が空で列1に値がある場合は詳細行
            is_section = sec_val in ['売上高', '売上原価', '販売費及び一般管理費', '営業外収益', '営業外費用', '特別利益', '特別損失', '営業外収益合計', '営業外費用合計', '特別利益合計', '特別損失合計', '当期純利益合計']
            is_total = '合計' in account_name or '利益' in account_name
            
            # 詳細行はインデントを追加
            display_name = f"  {account_name}" if is_detail else account_name
            
            # 勘定科目名
            ws.cell(row=row, column=col).value = display_name
            ws.cell(row=row, column=col).border = self.thin_border
            
            if is_section:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.section_fill
            elif is_total:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.total_fill
            
            col += 1
            
            # 各月のデータ
            for month in month_cols:
                # 当期金額
                current_value = row_data[month] if pd.notna(row_data[month]) else 0
                try:
                    current_value = float(current_value)
                except (ValueError, TypeError):
                    current_value = 0
                
                # 前期金額
                prev_value = None  # Noneは前期データなし、0は前期が0円
                if df_zenki is not None and section_col_zenki is not None and month in df_zenki.columns:
                    # 勘定科目名で前期データを検索（行タイプも一致させる）
                    for zenki_idx, zenki_data in df_zenki.iterrows():
                        zenki_sec = str(zenki_data[section_col_zenki]).strip() if pd.notna(zenki_data[section_col_zenki]) else ""
                        zenki_det = str(zenki_data[detail_col_zenki]).strip() if detail_col_zenki and pd.notna(zenki_data[detail_col_zenki]) else ""
                        zenki_is_detail = not zenki_sec and zenki_det  # 前期の行が詳細行かどうか
                        zenki_account = zenki_det if zenki_is_detail else zenki_sec
                        
                        # 当期と前期で同じ行タイプ（詳細行 or セクション/合計行）のみマッチ
                        if zenki_account == account_name and zenki_is_detail == is_detail:
                            prev_val = zenki_data[month]
                            try:
                                prev_value = float(prev_val) if pd.notna(prev_val) else 0
                            except (ValueError, TypeError):
                                prev_value = 0
                            break
                
                # 比率（売上高に対する比率）
                ratio = 0
                if month in sales_data and sales_data[month] != 0:
                    ratio = current_value / sales_data[month]
                
                # 前期比（前期データがある場合のみ計算）
                yoy = None
                if prev_value is not None and prev_value != 0:
                    yoy = (current_value - prev_value) / abs(prev_value)
                
                # 金額
                ws.cell(row=row, column=col).value = current_value
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 比率
                ws.cell(row=row, column=col).value = ratio
                ws.cell(row=row, column=col).number_format = '0.0%'
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 前期（データがない場合は空白）
                if prev_value is not None:
                    ws.cell(row=row, column=col).value = prev_value
                    ws.cell(row=row, column=col).number_format = '#,##0'
                else:
                    ws.cell(row=row, column=col).value = ""
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
                
                # 前期比
                if yoy is not None:
                    ws.cell(row=row, column=col).value = yoy
                    ws.cell(row=row, column=col).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=col).value = ""
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                if is_section:
                    ws.cell(row=row, column=col).fill = self.section_fill
                elif is_total:
                    ws.cell(row=row, column=col).fill = self.total_fill
                col += 1
            
            row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        for i in range(2, col):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    def _build_bs(self, wb: Workbook) -> None:
        """貸借対照表シートを作成（左右分割レイアウト）"""
        if self.bs_df is None:
            logger.warning("貸借対照表データがありません")
            return
        
        ws = wb.create_sheet(title="貸借対照表")
        
        # ページ設定（A4縦）
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        
        # タイトル
        row = 1
        ws.merge_cells(f'A{row}:N{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"貸借対照表 - {self.config.target_year}年{self.config.target_month}月末時点"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 2
        
        # データを資産の部と負債・純資産の部に分割
        df = self.bs_df.copy()
        section_col = df.columns[0] if len(df.columns) > 0 else None
        detail_col = df.columns[1] if len(df.columns) > 1 else None
        
        # データ列を特定（開始月、期間借方、期間貸方、終了月、構成比）
        data_cols = [col for col in df.columns[2:] if not pd.isna(col)]
        
        # 資産の部と負債・純資産の部を分割
        assets_rows = []
        liabilities_rows = []
        current_section = None
        
        for idx, row_data in df.iterrows():
            sec_val = str(row_data[section_col]).strip() if pd.notna(row_data[section_col]) else ""
            det_val = str(row_data[detail_col]).strip() if detail_col and pd.notna(row_data[detail_col]) else ""
            
            # セクション判定
            if sec_val == '資産の部':
                current_section = 'assets'
            elif sec_val == '負債の部':
                current_section = 'liabilities'
            elif sec_val == '純資産の部':
                current_section = 'equity'
            
            # 行データを分類
            account_name = det_val if det_val else sec_val
            is_detail = not sec_val and det_val
            
            row_info = {
                'account_name': account_name,
                'is_detail': is_detail,
                'is_section': sec_val in ['資産の部', '負債の部', '純資産の部', '流動資産', '固定資産', 
                                          '流動負債', '固定負債', '株主資本', '資本金', '資本剰余金', 
                                          '利益剰余金', '有形固定資産', '無形固定資産', '投資その他の資産',
                                          '現金及び預金', '売上債権', '有価証券', '棚卸資産', 'その他流動資産',
                                          '仕入債務', 'その他流動負債', '繰延資産', '諸口', '評価・換算差額等',
                                          '新株予約権'],
                'is_total': '合計' in account_name,
                'data': {col: row_data[col] for col in data_cols} if data_cols else {}
            }
            
            if current_section == 'assets' or sec_val == '資産の部合計':
                assets_rows.append(row_info)
            elif current_section in ['liabilities', 'equity'] or sec_val in ['負債の部合計', '純資産の部合計', '負債・純資産の部合計']:
                liabilities_rows.append(row_info)
        
        # ヘッダー行
        # 左側（資産の部）
        left_headers = ['勘定科目', '期首残高', '期末残高']
        # 右側（負債・純資産の部）
        right_headers = ['勘定科目', '期首残高', '期末残高']
        
        # 使用するデータ列を決定
        start_col_name = None
        end_col_name = None
        for col in data_cols:
            col_str = str(col).lower()
            if '開始' in col_str or '期首' in col_str:
                start_col_name = col
            elif '終了' in col_str or '期末' in col_str:
                end_col_name = col
        
        # ヘッダー出力
        header_row = row
        # 左側ヘッダー
        for i, h in enumerate(left_headers):
            cell = ws.cell(row=header_row, column=i+1)
            cell.value = h
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 中央の空白列
        ws.cell(row=header_row, column=4).value = ""
        
        # 右側ヘッダー
        right_start_col = 5
        for i, h in enumerate(right_headers):
            cell = ws.cell(row=header_row, column=right_start_col+i)
            cell.value = h
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # データ出力（左右を同時に出力）
        max_rows = max(len(assets_rows), len(liabilities_rows))
        
        for i in range(max_rows):
            current_row = row + i
            
            # 左側（資産の部）
            if i < len(assets_rows):
                asset = assets_rows[i]
                display_name = f"  {asset['account_name']}" if asset['is_detail'] else asset['account_name']
                
                # 勘定科目名
                cell = ws.cell(row=current_row, column=1)
                cell.value = display_name
                cell.border = self.thin_border
                if asset['is_section']:
                    cell.font = Font(bold=True)
                    cell.fill = self.section_fill
                elif asset['is_total']:
                    cell.font = Font(bold=True)
                    cell.fill = self.total_fill
                
                # 期首残高
                cell = ws.cell(row=current_row, column=2)
                if start_col_name and start_col_name in asset['data']:
                    val = asset['data'][start_col_name]
                    if pd.notna(val):
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0'
                        except (ValueError, TypeError):
                            cell.value = val
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='right')
                if asset['is_section']:
                    cell.fill = self.section_fill
                elif asset['is_total']:
                    cell.fill = self.total_fill
                
                # 期末残高
                cell = ws.cell(row=current_row, column=3)
                if end_col_name and end_col_name in asset['data']:
                    val = asset['data'][end_col_name]
                    if pd.notna(val):
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0'
                        except (ValueError, TypeError):
                            cell.value = val
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='right')
                if asset['is_section']:
                    cell.fill = self.section_fill
                elif asset['is_total']:
                    cell.fill = self.total_fill
            
            # 右側（負債・純資産の部）
            if i < len(liabilities_rows):
                liab = liabilities_rows[i]
                display_name = f"  {liab['account_name']}" if liab['is_detail'] else liab['account_name']
                
                # 勘定科目名
                cell = ws.cell(row=current_row, column=right_start_col)
                cell.value = display_name
                cell.border = self.thin_border
                if liab['is_section']:
                    cell.font = Font(bold=True)
                    cell.fill = self.section_fill
                elif liab['is_total']:
                    cell.font = Font(bold=True)
                    cell.fill = self.total_fill
                
                # 期首残高
                cell = ws.cell(row=current_row, column=right_start_col+1)
                if start_col_name and start_col_name in liab['data']:
                    val = liab['data'][start_col_name]
                    if pd.notna(val):
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0'
                        except (ValueError, TypeError):
                            cell.value = val
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='right')
                if liab['is_section']:
                    cell.fill = self.section_fill
                elif liab['is_total']:
                    cell.fill = self.total_fill
                
                # 期末残高
                cell = ws.cell(row=current_row, column=right_start_col+2)
                if end_col_name and end_col_name in liab['data']:
                    val = liab['data'][end_col_name]
                    if pd.notna(val):
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0'
                        except (ValueError, TypeError):
                            cell.value = val
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='right')
                if liab['is_section']:
                    cell.fill = self.section_fill
                elif liab['is_total']:
                    cell.fill = self.total_fill
        
        # 資産の部合計と負債・純資産の部合計を同じ行に配置するための調整
        # 資産の部合計の行を探す
        assets_total_row = None
        for i, asset in enumerate(assets_rows):
            if asset['account_name'] == '資産の部合計':
                assets_total_row = row + i
                break
        
        # 負債・純資産の部合計の行を探す
        liab_total_idx = None
        for i, liab in enumerate(liabilities_rows):
            if liab['account_name'] == '負債・純資産の部合計':
                liab_total_idx = i
                break
        
        # 負債・純資産の部合計を資産の部合計と同じ行に出力
        if assets_total_row and liab_total_idx is not None:
            liab = liabilities_rows[liab_total_idx]
            
            # 勘定科目名
            cell = ws.cell(row=assets_total_row, column=right_start_col)
            cell.value = liab['account_name']
            cell.border = self.thin_border
            cell.font = Font(bold=True)
            cell.fill = self.total_fill
            
            # 期首残高
            cell = ws.cell(row=assets_total_row, column=right_start_col+1)
            if start_col_name and start_col_name in liab['data']:
                val = liab['data'][start_col_name]
                if pd.notna(val):
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='right')
            cell.fill = self.total_fill
            
            # 期末残高
            cell = ws.cell(row=assets_total_row, column=right_start_col+2)
            if end_col_name and end_col_name in liab['data']:
                val = liab['data'][end_col_name]
                if pd.notna(val):
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='right')
            cell.fill = self.total_fill
        
        # 列幅調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 3  # 中央の空白列
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def _write_dataframe_to_sheet(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int = 1,
        is_bs: bool = False
    ) -> None:
        """DataFrameをシートに書き込み"""
        # ヘッダー行
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = column_name
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # データ行
        for row_idx, row_data in enumerate(df.itertuples(index=False), start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                
                # 値の設定
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0'
                else:
                    cell.value = str(value)
                
                cell.border = self.thin_border
                
                # 数値列は右寄せ
                if isinstance(value, (int, float)) and not pd.isna(value):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # セクション行のスタイリング（BSの場合）
                if is_bs and col_num == 1:
                    str_value = str(value) if not pd.isna(value) else ""
                    if str_value in ['資産の部', '負債の部', '純資産の部', '流動資産', '固定資産', '流動負債', '固定負債']:
                        for c in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = self.section_fill
                            ws.cell(row=row_idx, column=c).font = Font(bold=True)
        
        # 列幅の自動調整
        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(df.columns[col_num - 1]))
            
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df), min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # 日本語文字は幅が広いので調整
            adjusted_width = min(max_length * 1.5 + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ==================== データ取得メソッド ====================
    
    def _get_value_from_suii(self, account_name: str, column_name: str) -> Optional[float]:
        """PL月次推移から値を取得"""
        if self.pl_suii_df is None:
            return None
        
        try:
            # 勘定科目列を探す
            account_col = None
            for col in self.pl_suii_df.columns:
                if '勘定科目' in str(col) or col == self.pl_suii_df.columns[0]:
                    account_col = col
                    break
            
            if account_col is None:
                return None
            
            # 該当行を検索
            mask = self.pl_suii_df[account_col].astype(str).str.contains(account_name, na=False)
            if mask.any():
                row = self.pl_suii_df[mask].iloc[0]
                # 列を探す
                for col in self.pl_suii_df.columns:
                    if column_name in str(col):
                        value = row[col]
                        if pd.notna(value):
                            return float(value)
            return None
        except Exception as e:
            logger.warning(f"値の取得に失敗: {account_name}, {column_name}, {e}")
            return None
    
    def _get_current_month_sales(self) -> Optional[float]:
        """当月売上高を取得"""
        return self._get_value_from_suii('売上高', f'{self.config.target_month}月')
    
    def _get_cumulative_sales(self) -> Optional[float]:
        """累計売上高を取得"""
        return self._get_value_from_suii('売上高', '合計')
    
    def _get_current_month_gross_profit(self) -> Optional[float]:
        """当月粗利益を取得"""
        return self._get_value_from_suii('粗利益', f'{self.config.target_month}月')
    
    def _get_cumulative_gross_profit(self) -> Optional[float]:
        """累計粗利益を取得"""
        return self._get_value_from_suii('粗利益', '合計')
    
    def _get_current_month_operating_profit(self) -> Optional[float]:
        """当月営業利益を取得"""
        return self._get_value_from_suii('営業利益', f'{self.config.target_month}月')
    
    def _get_cumulative_operating_profit(self) -> Optional[float]:
        """累計営業利益を取得"""
        return self._get_value_from_suii('営業利益', '合計')
    
    def _get_gross_profit_rate(self) -> Optional[float]:
        """粗利益率を計算"""
        sales = self._get_cumulative_sales()
        gross_profit = self._get_cumulative_gross_profit()
        if sales and gross_profit and sales != 0:
            return (gross_profit / sales) * 100
        return None
    
    def _get_operating_profit_rate(self) -> Optional[float]:
        """営業利益率を計算"""
        sales = self._get_cumulative_sales()
        operating_profit = self._get_cumulative_operating_profit()
        if sales and operating_profit and sales != 0:
            return (operating_profit / sales) * 100
        return None
    
    def _get_f_rate(self) -> Optional[float]:
        """原価率を計算"""
        sales = self._get_cumulative_sales()
        cost = self._get_value_from_suii('売上原価', '合計')
        if sales and cost and sales != 0:
            return (cost / sales) * 100
        return None
    
    def _get_l_rate(self) -> Optional[float]:
        """人件費率を計算"""
        sales = self._get_cumulative_sales()
        labor = self._get_value_from_suii('人件費', '合計')
        if sales and labor and sales != 0:
            return (labor / sales) * 100
        return None
    
    def _get_fl_rate(self) -> Optional[float]:
        """FL率を計算"""
        f_rate = self._get_f_rate()
        l_rate = self._get_l_rate()
        if f_rate is not None and l_rate is not None:
            return f_rate + l_rate
        return None
    
    def _get_value_from_bs(self, account_name: str) -> Optional[float]:
        """BSから値を取得"""
        if self.bs_df is None:
            return None
        
        try:
            # 勘定科目列を探す
            account_col = None
            for col in self.bs_df.columns:
                if '勘定科目' in str(col):
                    account_col = col
                    break
            
            if account_col is None:
                # 2番目の列を使用
                account_col = self.bs_df.columns[1] if len(self.bs_df.columns) > 1 else self.bs_df.columns[0]
            
            # 該当行を検索
            mask = self.bs_df[account_col].astype(str).str.contains(account_name, na=False)
            if mask.any():
                row = self.bs_df[mask].iloc[0]
                # 期末残高列を探す
                for col in self.bs_df.columns:
                    if '期末残高' in str(col):
                        value = row[col]
                        if pd.notna(value):
                            return float(value)
            return None
        except Exception as e:
            logger.warning(f"BS値の取得に失敗: {account_name}, {e}")
            return None
    
    def _get_current_ratio(self) -> Optional[float]:
        """流動比率を計算"""
        current_assets = self._get_value_from_bs('流動資産合計')
        current_liabilities = self._get_value_from_bs('流動負債合計')
        if current_assets and current_liabilities and current_liabilities != 0:
            return (current_assets / current_liabilities) * 100
        return None
    
    def _get_equity_ratio(self) -> Optional[float]:
        """自己資本比率を計算"""
        equity = self._get_value_from_bs('純資産合計')
        total_assets = self._get_value_from_bs('資産合計')
        if equity and total_assets and total_assets != 0:
            return (equity / total_assets) * 100
        return None
