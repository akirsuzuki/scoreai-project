"""
居酒屋出店計画のエクスポート機能
"""
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from decimal import Decimal

from ..models import IzakayaPlan
from ..mixins import SelectedCompanyMixin
from ..services.izakaya_plan_service import IzakayaPlanService
from ..services.export_service import ExportService


class IzakayaPlanExportView(SelectedCompanyMixin, LoginRequiredMixin, View):
    """居酒屋出店計画のエクスポートビュー"""
    
    def get(self, request, plan_id, format_type='pdf'):
        """
        居酒屋出店計画をエクスポート
        
        Args:
            request: HTTPリクエスト
            plan_id: 計画ID
            format_type: エクスポート形式（pdf, excel）
        """
        plan = get_object_or_404(
            IzakayaPlan,
            id=plan_id,
            company=self.this_company,
            user=request.user
        )
        
        # 計算が未実行の場合は実行
        if plan.is_draft or not plan.monthly_revenue:
            try:
                IzakayaPlanService.calculate_all(plan)
                plan.refresh_from_db()
            except Exception as e:
                return JsonResponse({'error': f'計算中にエラーが発生しました: {str(e)}'}, status=400)
        
        if format_type == 'pdf':
            return self._export_pdf(plan)
        elif format_type == 'excel':
            return self._export_excel(plan)
        else:
            return JsonResponse({'error': 'サポートされていない形式です。'}, status=400)
    
    def _export_pdf(self, plan):
        """PDF形式でエクスポート"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            from io import BytesIO
            
            # 日本語フォントの登録
            try:
                font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static/fonts/ipaexg.ttf')
                pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                font_name = 'JapaneseFont'
            except:
                font_name = 'Helvetica'
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                leading=12
            )
            
            # 表紙
            story.append(Paragraph('居酒屋出店計画書', title_style))
            story.append(Spacer(1, 0.5 * inch))
            story.append(Paragraph(f'会社名: {plan.company.name}', normal_style))
            story.append(Paragraph(f'作成日: {datetime.now().strftime("%Y年%m月%d日")}', normal_style))
            story.append(PageBreak())
            
            # 事業概要
            story.append(Paragraph('1. 事業概要', title_style))
            story.append(Spacer(1, 0.2 * inch))
            
            overview_data = [
                ['項目', '内容'],
                ['店のコンセプト', plan.store_concept or '未設定'],
                ['席数', f'{plan.number_of_seats}席'],
                ['営業時間', f'{plan.opening_hours_start.strftime("%H:%M")} ～ {plan.opening_hours_end.strftime("%H:%M")}'],
                ['ターゲット顧客', plan.target_customer or '未設定'],
                ['客単価', f'{plan.average_price_per_customer:,.0f}円'],
            ]
            
            overview_table = Table(overview_data)
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(overview_table)
            story.append(PageBreak())
            
            # 投資計画
            story.append(Paragraph('2. 投資計画', title_style))
            story.append(Spacer(1, 0.2 * inch))
            
            investment_data = [
                ['項目', '金額'],
                ['初期投資額', f'{plan.initial_investment:,.0f}円'],
                ['月額家賃', f'{plan.monthly_rent:,.0f}円'],
                ['社員人数', f'{plan.number_of_staff}人'],
                ['社員月給合計', f'{plan.number_of_staff * plan.staff_monthly_salary:,.0f}円'],
                ['アルバイト時間数/月', f'{plan.part_time_hours_per_month}時間'],
                ['アルバイト時給', f'{plan.part_time_hourly_wage:,.0f}円'],
                ['アルバイト人件費/月', f'{plan.part_time_hours_per_month * plan.part_time_hourly_wage:,.0f}円'],
            ]
            
            investment_table = Table(investment_data)
            investment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(investment_table)
            story.append(PageBreak())
            
            # 収支計画
            story.append(Paragraph('3. 収支計画', title_style))
            story.append(Spacer(1, 0.2 * inch))
            
            # 月次収支表（12ヶ月分）
            monthly_data = [['月', '売上', '経費', '利益']]
            for month in range(1, 13):
                monthly_data.append([
                    f'{month}月',
                    f'{plan.monthly_revenue:,.0f}円',
                    f'{plan.monthly_cost:,.0f}円',
                    f'{plan.monthly_profit:,.0f}円'
                ])
            
            monthly_table = Table(monthly_data)
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(monthly_table)
            story.append(Spacer(1, 0.3 * inch))
            
            # 年次収支表（3年分）
            yearly_data = [['年度', '売上', '経費', '利益']]
            for year in range(1, 4):
                yearly_data.append([
                    f'{year}年目',
                    f'{plan.monthly_revenue * 12:,.0f}円',
                    f'{plan.monthly_cost * 12:,.0f}円',
                    f'{plan.monthly_profit * 12:,.0f}円'
                ])
            
            yearly_table = Table(yearly_data)
            yearly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(yearly_table)
            story.append(PageBreak())
            
            # 回収期間分析
            story.append(Paragraph('4. 回収期間分析', title_style))
            story.append(Spacer(1, 0.2 * inch))
            
            if plan.payback_period_years == 999:
                payback_text = '回収不可能（月間利益が0以下）'
            else:
                payback_text = f'{plan.payback_period_years}年{plan.payback_period_months}ヶ月'
            
            payback_data = [
                ['項目', '内容'],
                ['初期投資額', f'{plan.initial_investment:,.0f}円'],
                ['月間利益', f'{plan.monthly_profit:,.0f}円'],
                ['初期投資回収期間', payback_text],
            ]
            
            payback_table = Table(payback_data)
            payback_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(payback_table)
            
            # PDF生成
            doc.build(story)
            
            response = HttpResponse(content_type='application/pdf')
            filename = f"居酒屋出店計画書_{plan.company.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response.write(buffer.getvalue())
            buffer.close()
            
            return response
        except ImportError:
            return JsonResponse({'error': 'reportlabがインストールされていません。'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'PDF生成中にエラーが発生しました: {str(e)}'}, status=500)
    
    def _export_excel(self, plan):
        """Excel形式でエクスポート"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = '居酒屋出店計画'
            
            # スタイル定義
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=16)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row_num = 1
            
            # タイトル
            ws.merge_cells(f'A{row_num}:B{row_num}')
            title_cell = ws.cell(row=row_num, column=1)
            title_cell.value = '居酒屋出店計画書'
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 2
            
            # 基本情報
            ws.cell(row=row_num, column=1).value = '1. 事業概要'
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 1
            
            overview_data = [
                ['項目', '内容'],
                ['店のコンセプト', plan.store_concept or '未設定'],
                ['席数', f'{plan.number_of_seats}席'],
                ['営業時間', f'{plan.opening_hours_start.strftime("%H:%M")} ～ {plan.opening_hours_end.strftime("%H:%M")}'],
                ['ターゲット顧客', plan.target_customer or '未設定'],
                ['客単価', f'{plan.average_price_per_customer:,.0f}円'],
            ]
            
            for row_data in overview_data:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if row_num == row_num - len(overview_data) + 1:  # ヘッダー行
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row_num += 1
            
            row_num += 2
            
            # 投資計画
            ws.cell(row=row_num, column=1).value = '2. 投資計画'
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 1
            
            investment_data = [
                ['項目', '金額'],
                ['初期投資額', f'{plan.initial_investment:,.0f}円'],
                ['月額家賃', f'{plan.monthly_rent:,.0f}円'],
                ['社員人数', f'{plan.number_of_staff}人'],
                ['社員月給合計', f'{plan.number_of_staff * plan.staff_monthly_salary:,.0f}円'],
                ['アルバイト時間数/月', f'{plan.part_time_hours_per_month}時間'],
                ['アルバイト時給', f'{plan.part_time_hourly_wage:,.0f}円'],
                ['アルバイト人件費/月', f'{plan.part_time_hours_per_month * plan.part_time_hourly_wage:,.0f}円'],
            ]
            
            for row_data in investment_data:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if row_num == row_num - len(investment_data) + 1:  # ヘッダー行
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row_num += 1
            
            row_num += 2
            
            # 収支計画
            ws.cell(row=row_num, column=1).value = '3. 収支計画'
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 1
            
            # 月次収支表
            monthly_data = [['月', '売上', '経費', '利益']]
            for month in range(1, 13):
                monthly_data.append([
                    f'{month}月',
                    plan.monthly_revenue,
                    plan.monthly_cost,
                    plan.monthly_profit
                ])
            
            for row_data in monthly_data:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if row_num == row_num - len(monthly_data) + 1:  # ヘッダー行
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row_num += 1
            
            row_num += 2
            
            # 回収期間分析
            ws.cell(row=row_num, column=1).value = '4. 回収期間分析'
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 1
            
            if plan.payback_period_years == 999:
                payback_text = '回収不可能（月間利益が0以下）'
            else:
                payback_text = f'{plan.payback_period_years}年{plan.payback_period_months}ヶ月'
            
            payback_data = [
                ['項目', '内容'],
                ['初期投資額', plan.initial_investment],
                ['月間利益', plan.monthly_profit],
                ['初期投資回収期間', payback_text],
            ]
            
            for row_data in payback_data:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if row_num == row_num - len(payback_data) + 1:  # ヘッダー行
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row_num += 1
            
            # 列幅の自動調整
            for col_num in range(1, 3):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if row.value:
                            max_length = max(max_length, len(str(row.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"居酒屋出店計画書_{plan.company.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
        except ImportError:
            return JsonResponse({'error': 'openpyxlがインストールされていません。'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Excel生成中にエラーが発生しました: {str(e)}'}, status=500)

